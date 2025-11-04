#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
final_results_extract_ro_safe.py

Recursively scan for “final results” Excel files (any naming variation),
extract required fields from a fixed template-like sheet, and write a single JSON file.

This version is **read-only safe for merged cells**:
- It does **NOT** write to worksheets.
- It **virtually resolves** merged ranges by reading the top-left value for all cells
  covered by a merge region (so you can treat merged headers/values as repeated).

Outputs:
  - final_results_metadata.json (default) : JSON array of records
  - failed_files.txt (default)           : files that failed to parse (one absolute path per line)
  - run.log (default)                    : rotating human-readable log
  - events.jsonl (optional)              : structured event logs (JSONL), one per line

Usage example:
  python final_results_extract_ro_safe.py "D:/reports" \
      --out-json "D:/reports/final_results_metadata.json" \
      --failed-list "D:/reports/failed_files.txt" \
      --domains "D:/reports/domains.json" \
      --sheet-name "Summary" \
      --log-file "D:/reports/run.log"

domains.json format (recommended):
{
  "domains": ["FP_Modernization", "APP admin", "Fee Return"],
  "aliases": {
    "fp_mod": "FP_Modernization",
    "fp modernization": "FP_Modernization",
    "application admin": "APP admin",
    "fee_return": "Fee Return"
  }
}
"""

import argparse
import concurrent.futures as cf
import json
import logging
import os
import re
import sys
import traceback
from datetime import datetime
from logging.handlers import RotatingFileHandler
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# Third-party
try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
except Exception:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    raise

# -----------------------------
# Configuration defaults
# -----------------------------
FINAL_NAME_REGEX = re.compile(r"(?i).*final\W*results?.*\.(xlsx|xlsm)$")  # filename rule
HEADER_ROWS_TO_SNIFF = 3                      # rows to combine for header consolidation
MAX_ROWS_TO_SCAN_FOR_VALUES = 200             # scan window into data rows
RECOGNIZED_EXTS = {".xlsx", ".xlsm"}          # formats handled by openpyxl

# Normalized header candidates per field
REQ_DESC_HEADERS   = ["request_description"]
WORK_TRACK_HEADERS = ["work_track", "workstream", "work_stream", "track"]
RESULT_LOC_HEADERS = ["result_location", "sprint"]  # template might call it either
OUTPUT_LOGS_HEADERS = ["output_and_logs", "output_logs", "outputs_and_logs"]

# Regexes
RE_STORY = re.compile(r"\bUS\d{3,}\b", re.IGNORECASE)
RE_SPRINTS = [
    re.compile(r"\bPE\s*\d+(?:\.\d+)?\b", re.IGNORECASE),
    re.compile(r"\bPI\s*\d+(?:\.\d+)?\b", re.IGNORECASE),
    re.compile(r"\b(?:Sprint|S|SPR)\s*\d+\b", re.IGNORECASE),
]

WHITESPACE_RUN = re.compile(r"[ \t\r\f\v\u00A0\u200B]+")
NEWLINE_RUN    = re.compile(r"[\n\r]+")

# -----------------------------
# Logging
# -----------------------------
def setup_logging(log_file: Optional[str], jsonl_events: Optional[str]):
    logger = logging.getLogger("final_results")
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()

    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", "%Y-%m-%d %H:%M:%S")

    # Console (INFO+)
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)
    logger.addHandler(ch)

    # Rotating file (DEBUG+)
    if log_file:
        fh = RotatingFileHandler(log_file, maxBytes=5 * 1024 * 1024, backupCount=5, encoding="utf-8")
        fh.setLevel(logging.DEBUG)
        fh.setFormatter(fmt)
        logger.addHandler(fh)

    jsonl_handle = None
    if jsonl_events:
        jsonl_handle = open(jsonl_events, "w", encoding="utf-8")

    return logger, jsonl_handle

def log_event(jsonl_handle, event: Dict[str, Any]):
    if not jsonl_handle: return
    try:
        jsonl_handle.write(json.dumps(event, ensure_ascii=False) + "\n")
        jsonl_handle.flush()
    except Exception:
        pass

# -----------------------------
# Text / header helpers
# -----------------------------
def normalize_spaces(s: Any) -> str:
    if s is None:
        return ""
    s = str(s)
    s = s.replace("\u00A0", " ").replace("\u200B", "")  # NBSP, zero-width
    s = NEWLINE_RUN.sub(" ", s)
    s = WHITESPACE_RUN.sub(" ", s)
    return s.strip()

def normalize_header(s: str) -> str:
    s = normalize_spaces(s).lower()
    s = re.sub(r"[^0-9a-z]+", "_", s)
    s = re.sub(r"_+", "_", s)
    return s.strip("_")

def find_first_header(headers: List[str], candidates: List[str]) -> Optional[int]:
    norm = [normalize_header(h) for h in headers]
    cand_norm = set(candidates)
    for idx, h in enumerate(norm):
        if h in cand_norm:
            return idx
    return None

def filename_looks_like_final_results(path: Path) -> bool:
    return bool(FINAL_NAME_REGEX.match(path.name))

def to_abs(p: Path) -> str:
    try:
        return str(p.resolve())
    except Exception:
        return str(p.absolute())

# -----------------------------
# Domain resolution (optional)
# -----------------------------
class DomainResolver:
    def __init__(self, domains: List[str], aliases: Dict[str, str]):
        self.canon = {self._norm(d): d for d in domains}
        self.aliases = {self._norm(k): v for k, v in aliases.items()} if aliases else {}

    def _norm(self, s: str) -> str:
        s = normalize_spaces(s).lower()
        s = re.sub(r"[^0-9a-z]+", "_", s)
        s = re.sub(r"_+", "_", s)
        return s.strip("_")

    def resolve(self, text: str) -> Tuple[Optional[str], str]:
        if not text:
            return None, "unresolved"
        t = self._norm(text)
        # exact canonical
        if t in self.canon:
            return self.canon[t], "exact"
        # alias
        if t in self.aliases:
            return self.aliases[t], "alias"
        # conservative contains (canonical keys)
        for k_norm, dom in self.canon.items():
            if k_norm and k_norm in t:
                return dom, "contains"
        # alias contains
        for a_norm, dom in self.aliases.items():
            if a_norm and a_norm in t:
                return dom, "contains"
        return None, "unresolved"

def load_domain_config(path: Optional[str], logger: logging.Logger) -> DomainResolver:
    if not path:
        logger.warning("No domain file provided; business_domain may be unresolved.")
        return DomainResolver(domains=[], aliases={})
    p = Path(path)
    if not p.exists():
        logger.error(f"Domain file does not exist: {path}. Using empty domain set.")
        return DomainResolver(domains=[], aliases={})
    try:
        if p.suffix.lower() == ".json":
            data = json.loads(p.read_text(encoding="utf-8"))
            domains = data.get("domains", []) or []
            aliases = data.get("aliases", {}) or {}
            if not isinstance(domains, list) or not isinstance(aliases, dict):
                raise ValueError("Invalid domains.json structure")
            return DomainResolver(domains, aliases)
        else:
            domains = [line.strip() for line in p.read_text(encoding="utf-8").splitlines() if line.strip()]
            return DomainResolver(domains, aliases={})
    except Exception as e:
        logger.error(f"Failed to load domain config: {e}")
        return DomainResolver(domains=[], aliases={})

# -----------------------------
# Merged cells (virtual resolution)
# -----------------------------
def merged_regions(ws: Worksheet) -> List[Dict[str, int]]:
    """
    Snapshot merged regions and TL values. Requires a normal Worksheet
    (read_only=False). If attribute isn't available, returns [].
    """
    regs = []
    # Some worksheets may not expose merged_cells (rare in normal mode, common in read-only)
    if not hasattr(ws, "merged_cells"):
        return regs
    try:
        for r in list(ws.merged_cells.ranges):
            regs.append({
                "min_row": r.min_row, "min_col": r.min_col,
                "max_row": r.max_row, "max_col": r.max_col,
                "value": ws.cell(row=r.min_row, column=r.min_col).value
            })
    except Exception:
        # If any unexpected format, just return empty; we'll read raw cells.
        return []
    return regs

def sheet_to_2d_resolved(ws: Worksheet) -> List[List[Any]]:
    """
    Build a 2D list of values while resolving merged cells *virtually*:
    For any (r,c) inside a merged range, return the top-left value of that range.
    No writes to the worksheet, so it's safe with MergedCell objects.
    """
    regs = merged_regions(ws)
    max_row, max_col = ws.max_row, ws.max_column
    data: List[List[Any]] = []

    # Fast path if no merges
    if not regs:
        for r in range(1, max_row + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
            data.append(row_vals)
        return data

    # With merges, check membership; for performance this is O(R*C*merges) but header/data windows are small.
    for r in range(1, max_row + 1):
        row_vals = []
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                for M in regs:
                    if M["min_row"] <= r <= M["max_row"] and M["min_col"] <= c <= M["max_col"]:
                        v = M["value"]
                        break
            row_vals.append(v)
        data.append(row_vals)
    return data

# -----------------------------
# Header consolidation
# -----------------------------
def consolidate_headers(grid: List[List[Any]], header_rows: int) -> Tuple[List[str], int]:
    """
    Combine the first `header_rows` rows into a single header list of strings.
    Returns (headers, data_start_row_index).
    """
    if not grid:
        return [], 0
    rows = grid[:min(header_rows, len(grid))]
    max_cols = max((len(r) for r in rows), default=0)
    headers: List[str] = []
    for c in range(max_cols):
        parts = []
        for r in rows:
            if c < len(r):
                cell = r[c]
                if cell is None:
                    continue
                txt = normalize_spaces(str(cell))
                if txt:
                    parts.append(txt)
        headers.append(normalize_header(" ".join(parts)) if parts else f"col_{c+1}")
    return headers, len(rows)

# -----------------------------
# Field extraction
# -----------------------------
def extract_required_fields(
    headers: List[str],
    grid: List[List[Any]],
    data_start_idx: int,
    resolver: DomainResolver,
    logger: logging.Logger,
    file_path: str,
    sheet_name: str,
    strict_domain: bool = False,
) -> Dict[str, Any]:

    # Column indices
    req_desc_idx   = find_first_header(headers, REQ_DESC_HEADERS)
    work_track_idx = find_first_header(headers, WORK_TRACK_HEADERS)
    result_loc_idx = find_first_header(headers, RESULT_LOC_HEADERS)
    output_logs_idx = find_first_header(headers, OUTPUT_LOGS_HEADERS)

    # story_id
    story_id = None
    if req_desc_idx is not None:
        for r in range(data_start_idx, min(len(grid), data_start_idx + MAX_ROWS_TO_SCAN_FOR_VALUES)):
            row = grid[r]
            if req_desc_idx < len(row):
                cell = normalize_spaces(row[req_desc_idx])
                if not cell:
                    continue
                m = RE_STORY.search(cell)
                if m:
                    story_id = m.group(0).upper()
                    break

    # sprint + capture one useful text from Result Location (for domain too)
    sprint = None
    result_loc_text = ""
    if result_loc_idx is not None:
        for r in range(data_start_idx, min(len(grid), data_start_idx + MAX_ROWS_TO_SCAN_FOR_VALUES)):
            row = grid[r]
            if result_loc_idx < len(row):
                t = normalize_spaces(row[result_loc_idx])
                if t:
                    result_loc_text = t
                    for rx in RE_SPRINTS:
                        mm = rx.search(t)
                        if mm:
                            s = mm.group(0)
                            # normalize prefix spacing
                            s = re.sub(r"(?i)^(pe|pi|s|spr)\s*", lambda m: m.group(1).upper() + " ", s)
                            s = re.sub(r"\s+", " ", s).strip()
                            s = re.sub(r"(?i)^(spr)\s+(\d+)$", r"Sprint \2", s)
                            s = re.sub(r"(?i)^(s)\s+(\d+)$", r"Sprint \2", s)
                            sprint = s
                            break
                    break

    # work_track
    work_track = None
    if work_track_idx is not None:
        for r in range(data_start_idx, min(len(grid), data_start_idx + MAX_ROWS_TO_SCAN_FOR_VALUES)):
            row = grid[r]
            if work_track_idx < len(row):
                t = normalize_spaces(row[work_track_idx])
                if t:
                    work_track = t
                    break

    # business_domain (optional)
    business_domain = None
    domain_method = "unresolved"

    lookup_texts: List[str] = []
    if result_loc_text:
        lookup_texts.append(result_loc_text)

    if output_logs_idx is not None:
        for r in range(data_start_idx, min(len(grid), data_start_idx + MAX_ROWS_TO_SCAN_FOR_VALUES)):
            row = grid[r]
            if output_logs_idx < len(row):
                t = normalize_spaces(row[output_logs_idx])
                if t:
                    lookup_texts.append(t)
                    break

    for txt in lookup_texts:
        dom, method = resolver.resolve(txt)
        if dom:
            business_domain = dom
            domain_method = method
            break

    # validate required
    missing = []
    if not story_id:
        missing.append("story_id")
    if not sprint:
        missing.append("sprint")
    if strict_domain and not business_domain:
        missing.append("business_domain")

    return {
        "story_id": story_id,
        "work_track": work_track,
        "sprint": sprint,
        "business_domain": business_domain,
        "domain_method": domain_method,
        "missing": missing,
        "result_location_text": result_loc_text,  # optional audit
    }

# -----------------------------
# File processing
# -----------------------------
def process_file(
    path: Path,
    resolver: DomainResolver,
    sheet_name: Optional[str],
    strict_domain: bool,
    logger: logging.Logger,
    jsonl_handle: Optional[Any],
) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:

    file_path = to_abs(path)
    t0 = datetime.now()

    try:
        # IMPORTANT: open in normal mode so merged ranges metadata is available
        wb = load_workbook(filename=file_path, data_only=True, read_only=False, keep_links=False)
    except Exception as e:
        logger.error(f"Failed to open workbook: {file_path} :: {e}")
        log_event(jsonl_handle, {"event": "open_file", "status": "error", "file_path": file_path, "error": str(e)})
        return None, f"open_failed: {e}"

    try:
        # select sheet
        if sheet_name:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]  # type: ignore[assignment]
            else:
                wb.close()
                msg = f"sheet_not_found:{sheet_name}"
                logger.error(f"{msg} in {file_path}")
                log_event(jsonl_handle, {"event": "sheet_select", "status": "error", "file_path": file_path, "sheet": sheet_name, "error": "not_found"})
                return None, msg
        else:
            ws = wb[wb.sheetnames[0]]  # type: ignore[assignment]

        # Log merge info (no writing!)
        merge_count = 0
        if hasattr(ws, "merged_cells") and ws.merged_cells:
            try:
                merge_count = len(ws.merged_cells.ranges)  # type: ignore[attr-defined]
            except Exception:
                merge_count = 0
        logger.info(f"[merge_info] {file_path} : {merge_count} merged ranges (virtual resolution)")

        # Build grid with virtual merge resolution
        grid = sheet_to_2d_resolved(ws)
        headers, data_start_idx = consolidate_headers(grid, HEADER_ROWS_TO_SNIFF)

        log_event(jsonl_handle, {
            "event": "headers_built", "status": "success",
            "file_path": file_path, "sheet_name": ws.title,
            "headers": headers, "data_start_idx": data_start_idx
        })

        # Extract fields
        fields = extract_required_fields(
            headers=headers,
            grid=grid,
            data_start_idx=data_start_idx,
            resolver=resolver,
            logger=logger,
            file_path=file_path,
            sheet_name=ws.title,
            strict_domain=strict_domain,
        )

        missing = fields.get("missing", [])
        if missing:
            wb.close()
            logger.error(f"[required_missing] {file_path} -> missing: {missing}")
            log_event(jsonl_handle, {"event": "extract_fields", "status": "error", "file_path": file_path, "missing": missing})
            return None, f"required_missing:{','.join(missing)}"

        record = {
            "file_path": file_path,
            "story_id": fields["story_id"],
            "work_track": fields["work_track"],
            "sprint": fields["sprint"],
            "business_domain": fields["business_domain"],  # may be None if no --domains provided
        }

        wb.close()
        elapsed_ms = int((datetime.now() - t0).total_seconds() * 1000)
        logger.info(f"[processed] {file_path} in {elapsed_ms} ms")
        log_event(jsonl_handle, {"event": "file_processed", "status": "success", "file_path": file_path, "elapsed_ms": elapsed_ms, "record_preview": record})
        return record, None

    except Exception as e:
        tb = traceback.format_exc(limit=2)
        logger.error(f"[process_error] {file_path} :: {e}")
        log_event(jsonl_handle, {"event": "process_error", "status": "error", "file_path": file_path, "error": str(e), "trace": tb})
        try:
            wb.close()
        except Exception:
            pass
        return None, f"process_error:{e}"

# -----------------------------
# Discovery
# -----------------------------
def discover_files(root: Path, logger: logging.Logger) -> List[Path]:
    files: List[Path] = []
    for base, _, filenames in os.walk(root):
        for name in filenames:
            p = Path(base) / name
            if p.suffix.lower() in RECOGNIZED_EXTS and filename_looks_like_final_results(p):
                files.append(p)
    files.sort(key=lambda x: str(x).lower())
    logger.info(f"[discovery] found {len(files)} candidate files under {to_abs(root)}")
    return files

# -----------------------------
# Main
# -----------------------------
def main():
    ap = argparse.ArgumentParser(description="Extract metadata from 'final results' Excel files into JSON (RO-safe merges).")
    ap.add_argument("root", help="Root directory to scan recursively")
    ap.add_argument("--out-json", default="final_results_metadata.json", help="Output JSON file (array)")
    ap.add_argument("--failed-list", default="failed_files.txt", help="Path to write list of failed files")
    ap.add_argument("--domains", default=None, help="Domain config file (JSON recommended, or TXT one-per-line)")
    ap.add_argument("--log-file", default="run.log", help="Path to rotating log file")
    ap.add_argument("--events-jsonl", default=None, help="Optional JSONL structured events log")
    ap.add_argument("--sheet-name", default=None, help="Exact sheet name if fixed by template (default: first sheet)")
    ap.add_argument("--strict-domain", action="store_true", help="Treat unresolved business_domain as a hard failure")
    ap.add_argument("--workers", type=int, default=4, help="Thread pool size")
    args = ap.parse_args()

    logger, jsonl_handle = setup_logging(args.log_file, args.events_jsonl)
    resolver = load_domain_config(args.domains, logger)

    root = Path(args.root)
    if not root.exists():
        logger.error(f"Root path does not exist: {args.root}")
        sys.exit(2)

    files = discover_files(root, logger)
    results: List[Dict[str, Any]] = []
    failed: List[Tuple[str, str]] = []

    with cf.ThreadPoolExecutor(max_workers=max(1, args.workers)) as ex:
        futs = [ex.submit(process_file, p, resolver, args.sheet_name, args.strict_domain, logger, jsonl_handle) for p in files]
        for fut, p in zip(futs, files):
            rec, err = fut.result()
            if rec:
                results.append(rec)
            else:
                failed.append((to_abs(p), err or "unknown_error"))

    # Write outputs
    try:
        Path(args.out_json).write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")
        logger.info(f"[output_json] Wrote {len(results)} records to {args.out_json}")
        log_event(jsonl_handle, {"event": "write_output", "status": "success", "out_json": args.out_json, "records": len(results)})
    except Exception as e:
        logger.error(f"Failed to write output JSON: {e}")
        log_event(jsonl_handle, {"event": "write_output", "status": "error", "error": str(e)})

    # Failed list (unique sorted)
    try:
        unique_sorted = sorted({p for p, _ in failed})
        Path(args.failed_list).write_text("\n".join(unique_sorted), encoding="utf-8")
        logger.info(f"[failed_list] Wrote {len(unique_sorted)} failed files to {args.failed_list}")
        if failed:
            # optional reasons file (same stem, .jsonl)
            reasons_path = str(Path(args.failed_list).with_suffix(".jsonl"))
            with open(reasons_path, "w", encoding="utf-8") as fh:
                for p, reason in failed:
                    fh.write(json.dumps({"file_path": p, "reason": reason}, ensure_ascii=False) + "\n")
            logger.info(f"[failed_reasons] Wrote reasons to {reasons_path}")
    except Exception as e:
        logger.error(f"Failed to write failed files list: {e}")

    # Summary
    logger.info(
        "[summary] discovered=%d ok=%d failed=%d out=%s failed_list=%s",
        len(files), len(results), len({p for p, _ in failed}), args.out_json, args.failed_list
    )

    if jsonl_handle:
        try:
            jsonl_handle.close()
        except Exception:
            pass

if __name__ == "__main__":
    main()