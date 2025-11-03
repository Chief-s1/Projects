#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
final_results_extract.py - OPTIMIZED VERSION

Key improvements:
1. Read-only mode for faster workbook loading
2. Lazy header consolidation (only when needed)
3. Cached regex compilation
4. Early exit strategies to reduce scanning
5. Memory-efficient grid processing
6. Better error handling and resource management
7. Progress bar support
8. Smart domain matching with caching
"""

import argparse
import concurrent.futures as cf
import json
import logging
import os
import re
import sys
import traceback
from collections import defaultdict
from datetime import datetime
from functools import lru_cache
from logging.handlers import RotatingFileHandler
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError as e:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# Optional progress bar
try:
    from tqdm import tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False

# -----------------------------
# Configuration
# -----------------------------
FINAL_NAME_REGEX = re.compile(r"(?i).*final\W*results?.*\.(xlsx|xlsm)$")
HEADER_ROWS_TO_SNIFF = 3
MAX_ROWS_TO_SCAN_FOR_VALUES = 200
RECOGNIZED_EXTS = {".xlsx", ".xlsm"}

REQ_DESC_HEADERS = ["request_description"]
WORK_TRACK_HEADERS = ["work_track", "workstream", "work_stream", "track"]
RESULT_LOC_HEADERS = ["result_location", "sprint"]
OUTPUT_LOGS_HEADERS = ["output_and_logs", "output_logs", "outputs_and_logs"]

# Pre-compiled regexes
RE_STORY = re.compile(r"\bUS\d{3,}\b", re.IGNORECASE)
RE_SPRINTS = [
    re.compile(r"\bPE\s*\d+(?:\.\d+)?\b", re.IGNORECASE),
    re.compile(r"\bPI\s*\d+(?:\.\d+)?\b", re.IGNORECASE),
    re.compile(r"\b(?:Sprint|S|SPR)\s*\d+\b", re.IGNORECASE),
]
WHITESPACE_RUN = re.compile(r"[ \t\r\f\v\u00A0\u200B]+")
NEWLINE_RUN = re.compile(r"[\n\r]+")
SPRINT_NORMALIZE = re.compile(r"(?i)^(pe|pi|s|spr)\s*")
SPRINT_SPR = re.compile(r"(?i)^(spr)\s+(\d+)$")
SPRINT_S = re.compile(r"(?i)^(s)\s+(\d+)$")

# -----------------------------
# Logging
# -----------------------------
def setup_logging(log_file: Optional[str], jsonl_events: Optional[str]) -> Tuple[logging.Logger, Optional[Any]]:
    logger = logging.getLogger("final_results")
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()

    formatter = logging.Formatter(
        fmt="%(asctime)s [%(levelname)s] %(message)s", datefmt="%Y-%m-%d %H:%M:%S"
    )

    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(formatter)
    logger.addHandler(ch)

    if log_file:
        fh = RotatingFileHandler(log_file, maxBytes=5 * 1024 * 1024, backupCount=3, encoding="utf-8")
        fh.setLevel(logging.DEBUG)
        fh.setFormatter(formatter)
        logger.addHandler(fh)

    jsonl_handle = None
    if jsonl_events:
        jsonl_handle = open(jsonl_events, "w", encoding="utf-8", buffering=8192)

    return logger, jsonl_handle


def log_event(jsonl_handle, event: Dict[str, Any]):
    if jsonl_handle:
        try:
            jsonl_handle.write(json.dumps(event, ensure_ascii=False) + "\n")
        except Exception:
            pass

# -----------------------------
# Text processing with caching
# -----------------------------
@lru_cache(maxsize=2048)
def normalize_spaces(s: str) -> str:
    if not s:
        return ""
    s = str(s)
    s = s.replace("\u00A0", " ").replace("\u200B", "")
    s = NEWLINE_RUN.sub(" ", s)
    s = WHITESPACE_RUN.sub(" ", s)
    return s.strip()

@lru_cache(maxsize=512)
def normalize_header(s: str) -> str:
    s = normalize_spaces(s)
    s = s.lower()
    s = re.sub(r"[^0-9a-z]+", "_", s)
    s = re.sub(r"_+", "_", s)
    return s.strip("_")

def find_first_header(headers: List[str], candidates: List[str]) -> Optional[int]:
    """Optimized with set lookup"""
    cand_set = set(candidates)
    for idx, h in enumerate(headers):
        if normalize_header(h) in cand_set:
            return idx
    return None

def filename_looks_like_final_results(path: Path) -> bool:
    return bool(FINAL_NAME_REGEX.match(path.name))

def to_abs(p: Path) -> str:
    try:
        return str(p.resolve())
    except Exception:
        return str(p.absolute())

# -----------------------------
# Domain resolution with caching
# -----------------------------
class DomainResolver:
    def __init__(self, domains: List[str], aliases: Dict[str, str]):
        self.canon = {self._norm(d): d for d in domains}
        self.aliases = {}
        for k, v in aliases.items():
            self.aliases[self._norm(k)] = v
        # Cache for resolved domains
        self._cache = {}

    @staticmethod
    @lru_cache(maxsize=1024)
    def _norm(s: str) -> str:
        s = normalize_spaces(s)
        s = s.lower()
        s = re.sub(r"[^0-9a-z]+", "_", s)
        s = re.sub(r"_+", "_", s)
        return s.strip("_")

    def resolve(self, text: str) -> Tuple[Optional[str], str]:
        if not text:
            return None, "unresolved"
        
        # Check cache first
        if text in self._cache:
            return self._cache[text]
        
        t = self._norm(text)

        # Exact match
        if t in self.canon:
            result = (self.canon[t], "exact")
            self._cache[text] = result
            return result
        
        # Alias match
        if t in self.aliases:
            result = (self.aliases[t], "alias")
            self._cache[text] = result
            return result

        # Contains match (canonical)
        for key_norm, dom in self.canon.items():
            if key_norm and key_norm in t:
                result = (dom, "contains")
                self._cache[text] = result
                return result

        # Contains match (aliases)
        for alias_norm, dom in self.aliases.items():
            if alias_norm and alias_norm in t:
                result = (dom, "contains")
                self._cache[text] = result
                return result

        result = (None, "unresolved")
        self._cache[text] = result
        return result

def load_domain_config(path: Optional[str], logger: logging.Logger) -> DomainResolver:
    if not path:
        logger.warning("No domain file provided; business_domain may be unresolved.")
        return DomainResolver(domains=[], aliases={})
    
    p = Path(path)
    if not p.exists():
        logger.error(f"Domain file does not exist: {path}")
        return DomainResolver(domains=[], aliases={})

    try:
        if p.suffix.lower() == ".json":
            data = json.loads(p.read_text(encoding="utf-8"))
            domains = data.get("domains", []) or []
            aliases = data.get("aliases", {}) or {}
            return DomainResolver(domains=domains, aliases=aliases)
        else:
            domains = [line.strip() for line in p.read_text(encoding="utf-8").splitlines() if line.strip()]
            return DomainResolver(domains=domains, aliases={})
    except Exception as e:
        logger.error(f"Failed to load domain config: {e}")
        return DomainResolver(domains=[], aliases={})

# -----------------------------
# Excel processing - OPTIMIZED
# -----------------------------
def get_merged_value(ws: Worksheet, row: int, col: int, merged_ranges_cache: Dict) -> Any:
    """Get value from potentially merged cell using cache"""
    cell_coord = (row, col)
    if cell_coord in merged_ranges_cache:
        source_row, source_col = merged_ranges_cache[cell_coord]
        return ws.cell(row=source_row, column=source_col).value
    return ws.cell(row=row, column=col).value

def build_merged_cache(ws: Worksheet) -> Dict[Tuple[int, int], Tuple[int, int]]:
    """Build lookup cache for merged cells - faster than expanding all"""
    cache = {}
    for mcr in ws.merged_cells.ranges:
        source = (mcr.min_row, mcr.min_col)
        for r in range(mcr.min_row, mcr.max_row + 1):
            for c in range(mcr.min_col, mcr.max_col + 1):
                if (r, c) != source:
                    cache[(r, c)] = source
    return cache

def get_headers_lazy(ws: Worksheet, merged_cache: Dict, header_rows: int = 3) -> Tuple[List[str], int]:
    """Lazily build headers without loading entire sheet"""
    max_col = ws.max_column
    headers = []
    
    for c in range(1, max_col + 1):
        parts = []
        for r in range(1, min(header_rows + 1, ws.max_row + 1)):
            val = get_merged_value(ws, r, c, merged_cache)
            if val:
                txt = normalize_spaces(str(val))
                if txt:
                    parts.append(txt)
        
        if parts:
            headers.append(normalize_header(" ".join(parts)))
        else:
            headers.append(f"col_{c}")
    
    return headers, header_rows

def extract_required_fields_lazy(
    ws: Worksheet,
    headers: List[str],
    data_start_row: int,
    merged_cache: Dict,
    resolver: DomainResolver,
    logger: logging.Logger,
    file_path: str,
    strict_domain: bool = False,
) -> Dict[str, Any]:
    """Extract fields without loading entire sheet into memory"""
    
    # Find column indices
    req_desc_idx = find_first_header(headers, REQ_DESC_HEADERS)
    work_track_idx = find_first_header(headers, WORK_TRACK_HEADERS)
    result_loc_idx = find_first_header(headers, RESULT_LOC_HEADERS)
    output_logs_idx = find_first_header(headers, OUTPUT_LOGS_HEADERS)

    # Track what we've found for early exit
    story_id = None
    sprint = None
    work_track = None
    result_loc_text = ""
    output_logs_text = ""
    
    max_row = min(ws.max_row, data_start_row + MAX_ROWS_TO_SCAN_FOR_VALUES)
    
    # Single pass through rows
    for r in range(data_start_row + 1, max_row + 1):
        # Early exit if we have everything critical
        if story_id and sprint and work_track:
            break
        
        # Story ID
        if not story_id and req_desc_idx:
            val = get_merged_value(ws, r, req_desc_idx + 1, merged_cache)
            if val:
                cell = normalize_spaces(str(val))
                m = RE_STORY.search(cell)
                if m:
                    story_id = m.group(0).upper()
        
        # Sprint and result location text
        if not sprint and result_loc_idx:
            val = get_merged_value(ws, r, result_loc_idx + 1, merged_cache)
            if val:
                t = normalize_spaces(str(val))
                if t and not result_loc_text:
                    result_loc_text = t
                    for rx in RE_SPRINTS:
                        mm = rx.search(t)
                        if mm:
                            s = mm.group(0)
                            s = SPRINT_NORMALIZE.sub(lambda m: m.group(1).upper() + " ", s)
                            s = re.sub(r"\s+", " ", s).strip()
                            s = SPRINT_SPR.sub(r"Sprint \2", s)
                            s = SPRINT_S.sub(r"Sprint \2", s)
                            sprint = s
                            break
        
        # Work track
        if not work_track and work_track_idx:
            val = get_merged_value(ws, r, work_track_idx + 1, merged_cache)
            if val:
                t = normalize_spaces(str(val))
                if t:
                    work_track = t
        
        # Output logs (only get first occurrence)
        if not output_logs_text and output_logs_idx:
            val = get_merged_value(ws, r, output_logs_idx + 1, merged_cache)
            if val:
                t = normalize_spaces(str(val))
                if t:
                    output_logs_text = t

    # Business domain resolution
    business_domain = None
    domain_method = "unresolved"
    
    for txt in [result_loc_text, output_logs_text]:
        if txt:
            dom, method = resolver.resolve(txt)
            if dom:
                business_domain = dom
                domain_method = method
                break

    # Validate
    missing = []
    if not story_id:
        missing.append("story_id")
    if not sprint:
        missing.append("sprint")
    if strict_domain and not business_domain:
        missing.append("business_domain")

    return {
        "story_id": story_id,
        "work_track": work_track,
        "sprint": sprint,
        "business_domain": business_domain,
        "domain_method": domain_method,
        "missing": missing,
    }

# -----------------------------
# File processing - OPTIMIZED
# -----------------------------
def process_file(
    path: Path,
    resolver: DomainResolver,
    sheet_name: Optional[str],
    strict_domain: bool,
    logger: logging.Logger,
    jsonl_handle: Optional[Any],
) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
    """Process file using read-only mode and lazy evaluation"""
    file_path = to_abs(path)
    t0 = datetime.now()
    wb = None

    try:
        # Use read_only=True and data_only=True for faster loading
        wb = load_workbook(filename=file_path, data_only=True, read_only=True)
        
        # Select sheet
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                msg = f"sheet_not_found: {sheet_name}"
                logger.error(f"{msg} in {file_path}")
                log_event(jsonl_handle, {
                    "event": "sheet_select", "status": "error",
                    "file_path": file_path, "sheet": sheet_name, "error": "not_found"
                })
                return None, msg
            ws = wb[sheet_name]
        else:
            ws = wb[wb.sheetnames[0]]

        # Build merged cell cache (faster than expanding)
        merged_cache = build_merged_cache(ws)
        
        # Get headers lazily
        headers, data_start_row = get_headers_lazy(ws, merged_cache, HEADER_ROWS_TO_SNIFF)
        
        # Extract fields using lazy evaluation
        fields = extract_required_fields_lazy(
            ws=ws,
            headers=headers,
            data_start_row=data_start_row,
            merged_cache=merged_cache,
            resolver=resolver,
            logger=logger,
            file_path=file_path,
            strict_domain=strict_domain,
        )

        missing = fields.get("missing", [])
        if missing:
            logger.error(f"[required_missing] {file_path} -> missing: {missing}")
            log_event(jsonl_handle, {
                "event": "extract_fields", "status": "error",
                "file_path": file_path, "missing": missing
            })
            return None, f"required_missing:{','.join(missing)}"

        record = {
            "file_path": file_path,
            "story_id": fields["story_id"],
            "work_track": fields["work_track"],
            "sprint": fields["sprint"],
            "business_domain": fields["business_domain"],
        }

        elapsed_ms = int((datetime.now() - t0).total_seconds() * 1000)
        logger.info(f"[processed] {file_path} in {elapsed_ms}ms")
        log_event(jsonl_handle, {
            "event": "file_processed", "status": "success",
            "file_path": file_path, "elapsed_ms": elapsed_ms,
            "record_preview": record
        })
        return record, None

    except Exception as e:
        tb = traceback.format_exc(limit=3)
        logger.error(f"[process_error] {file_path} :: {e}")
        log_event(jsonl_handle, {
            "event": "process_error", "status": "error",
            "file_path": file_path, "error": str(e), "trace": tb
        })
        return None, f"process_error: {e}"
    
    finally:
        if wb:
            try:
                wb.close()
            except Exception:
                pass

# -----------------------------
# Discovery - OPTIMIZED
# -----------------------------
def discover_files(root: Path, logger: logging.Logger, skip_dirs: set = None) -> List[Path]:
    """Optimized file discovery with directory skipping"""
    if skip_dirs is None:
        skip_dirs = {".git", "node_modules", "__pycache__", ".venv", "venv"}
    
    files = []
    for base, dirs, filenames in os.walk(root):
        # Skip unwanted directories
        dirs[:] = [d for d in dirs if d not in skip_dirs]
        
        for name in filenames:
            if name.startswith('~$'):  # Skip temp Excel files
                continue
            
            p = Path(base) / name
            if p.suffix.lower() in RECOGNIZED_EXTS and filename_looks_like_final_results(p):
                files.append(p)
    
    files.sort(key=lambda x: str(x).lower())
    logger.info(f"[discovery] found {len(files)} candidate files under {to_abs(root)}")
    return files

# -----------------------------
# Main - OPTIMIZED
# -----------------------------
def main():
    ap = argparse.ArgumentParser(
        description="Extract metadata from 'final results' Excel files into JSON (OPTIMIZED)",
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument("root", help="Root directory to scan recursively")
    ap.add_argument("--out-json", default="final_results_metadata.json", help="Output JSON file")
    ap.add_argument("--failed-list", default="failed_files.txt", help="Failed files list")
    ap.add_argument("--domains", default=None, help="Domain config file (JSON)")
    ap.add_argument("--log-file", default="run.log", help="Log file path")
    ap.add_argument("--events-jsonl", default=None, help="JSONL events log")
    ap.add_argument("--sheet-name", default=None, help="Specific sheet name")
    ap.add_argument("--strict-domain", action="store_true", help="Require business_domain")
    ap.add_argument("--workers", type=int, default=4, help="Concurrent workers")
    ap.add_argument("--no-progress", action="store_true", help="Disable progress bar")
    args = ap.parse_args()

    logger, jsonl_handle = setup_logging(args.log_file, args.events_jsonl)
    resolver = load_domain_config(args.domains, logger)

    root = Path(args.root)
    if not root.exists():
        logger.error(f"Root path does not exist: {args.root}")
        sys.exit(2)

    files = discover_files(root, logger)
    if not files:
        logger.warning("No files found to process")
        sys.exit(0)

    results: List[Dict[str, Any]] = []
    failed: List[Tuple[str, str]] = []

    # Progress tracking
    show_progress = HAS_TQDM and not args.no_progress
    
    # Process files concurrently
    with cf.ThreadPoolExecutor(max_workers=max(1, args.workers)) as ex:
        futures = {ex.submit(
            process_file, p, resolver, args.sheet_name, 
            args.strict_domain, logger, jsonl_handle
        ): p for p in files}
        
        iterator = cf.as_completed(futures)
        if show_progress:
            iterator = tqdm(iterator, total=len(files), desc="Processing")
        
        for fut in iterator:
            p = futures[fut]
            try:
                rec, err = fut.result()
                if rec:
                    results.append(rec)
                else:
                    failed.append((to_abs(p), err or "unknown_error"))
            except Exception as e:
                logger.error(f"Future exception for {p}: {e}")
                failed.append((to_abs(p), f"future_exception: {e}"))

    # Write outputs
    try:
        with open(args.out_json, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        logger.info(f"[output_json] Wrote {len(results)} records to {args.out_json}")
        log_event(jsonl_handle, {
            "event": "write_output", "status": "success",
            "out_json": args.out_json, "records": len(results)
        })
    except Exception as e:
        logger.error(f"Failed to write output JSON: {e}")
        log_event(jsonl_handle, {
            "event": "write_output", "status": "error", "error": str(e)
        })

    # Write failed files
    try:
        unique_failed = sorted({p for p, _ in failed})
        Path(args.failed_list).write_text("\n".join(unique_failed), encoding="utf-8")
        logger.info(f"[failed_list] Wrote {len(unique_failed)} failed files to {args.failed_list}")
        
        if failed:
            reasons_path = str(Path(args.failed_list).with_suffix(".jsonl"))
            with open(reasons_path, "w", encoding="utf-8") as fh:
                for p, reason in failed:
                    fh.write(json.dumps({"file_path": p, "reason": reason}, ensure_ascii=False) + "\n")
            logger.info(f"[failed_reasons] Wrote reasons to {reasons_path}")
    except Exception as e:
        logger.error(f"Failed to write failed files list: {e}")

    # Summary
    logger.info(
        "[summary] discovered=%d ok=%d failed=%d out=%s",
        len(files), len(results), len({p for p, _ in failed}), args.out_json
    )

    if jsonl_handle:
        try:
            jsonl_handle.close()
        except Exception:
            pass


if __name__ == "__main__":
    main()