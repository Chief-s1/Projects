from openpyxl import load_workbook

# --- CONFIG ---
file_path = r"C:\path\to\your\excel_file.xlsx"   # your Excel file
sheet_name = "Sheet1"                            # change if needed
messages = ["msg1", "msg2", "msg3", "msg4"]      # repeat as many as you want
chunk_size = 100                                 # number of rows per message
column = "B"                                     # column to modify
start_row = 2                                    # skip header if needed (use 1 if not)

# --- SCRIPT ---
wb = load_workbook(file_path)
ws = wb[sheet_name]

row = start_row
for msg in messages:
    for _ in range(chunk_size):
        ws[f"{column}{row}"] = msg
        row += 1

# Optional: if total rows < chunks*100, it’ll just stop when Excel ends
wb.save(file_path)
wb.close()

print(f"✅ Updated column {column} with messages in chunks of {chunk_size}")