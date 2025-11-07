import os, re, json, logging, io
from logging.handlers import RotatingFileHandler
from pathlib import Path
from openpyxl import load_workbook

# ==== CONFIG ====
FOLDERS_FILE = Path(r"C:\Users\Swayam\Desktop\folders.txt")
OUT_JSONL    = Path("all_folders_story_ids.jsonl")
FAILED_LIST  = Path("failed_files.txt")
LOG_FILE     = Path("run.log")
PASSWORD     = "pe"

# ==== PATTERNS ====
RE_STORY = re.compile(r"\bUS\d{4,5}\b", re.IGNORECASE)

# ==== LOGGING ====
def setup_logger():
    lg = logging.getLogger("story_scan")
    lg.setLevel(logging.DEBUG)
    lg.handlers.clear()
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", "%Y-%m-%d %H:%M:%S")
    ch = logging.StreamHandler(); ch.setLevel(logging.INFO); ch.setFormatter(fmt); lg.addHandler(ch)
    fh = RotatingFileHandler(LOG_FILE, maxBytes=5*1024*1024, backupCount=5, encoding="utf-8")
    fh.setLevel(logging.DEBUG); fh.setFormatter(fmt); lg.addHandler(fh)
    return lg

logger = setup_logger()

# ==== HELPERS ====
def normalize_spaces(s):
    if s is None: return ""
    return str(s).replace("\u00A0"," ").replace("\u200B","").strip()

def merged_regions(ws):
    regs=[]
    if not hasattr(ws,"merged_cells"): return regs
    try:
        for r in list(ws.merged_cells.ranges):
            regs.append({
                "min_row": r.min_row, "min_col": r.min_col,
                "max_row": r.max_row, "max_col": r.max_col,
                "value": ws.cell(row=r.min_row, column=r.min_col).value
            })
    except Exception as e:
        logger.debug(f"[merge_regions_error] {ws.title}: {e}")
        return []
    return regs

def sheet_to_2d_resolved(ws):
    regs = merged_regions(ws)
    R, C = ws.max_row, ws.max_column
    out=[]
    if not regs:
        for r in range(1,R+1):
            out.append([ws.cell(row=r,column=c).value for c in range(1,C+1)])
        return out
    for r in range(1,R+1):
        row=[]
        for c in range(1,C+1):
            v = ws.cell(row=r,column=c).value
            if v is None:
                for M in regs:
                    if M["min_row"]<=r<=M["max_row"] and M["min_col"]<=c<=M["max_col"]:
                        v = M["value"]; break
            row.append(v)
        out.append(row)
    return out

def file_name_matches_final_results(file_path: Path) -> bool:
    stem = file_path.stem
    s = re.sub(r'([a-z])([A-Z])', r'\1 \2', stem)
    s = s.replace('_',' ').replace('-',' ')
    s = re.sub(r'[^0-9a-zA-Z]+',' ', s).lower()
    s = re.sub(r'\s+',' ', s).strip()
    tokens = s.split()
    has_final  = 'final' in tokens
    has_result = any(t in ('result','results') for t in tokens)
    phrase_fwd = re.search(r'\bfinal\s*result(s)?\b', s) is not None
    phrase_rev = re.search(r'\bresult(s)?\s*final\b', s) is not None
    return (has_final and has_result) or phrase_fwd or phrase_rev

# ==== OPEN WORKBOOK (with password fallback) ====
def open_workbook_any(p: Path):
    try:
        logger.debug(f"[open_workbook] {p}")
        wb = load_workbook(filename=str(p), data_only=True, read_only=False, keep_links=False)
        return wb, "plain"
    except PermissionError:
        logger.warning(f"[skip_locked] Permission denied: {p}")
        return None, "permission_denied"
    except Exception as e_plain:
        try:
            import msoffcrypto
        except Exception:
            logger.warning(f"[decrypt_unavailable] Install msoffcrypto-tool to open encrypted files: {p}")
            return None, f"open_error:{e_plain}"
        try:
            if p.suffix.lower() == ".xls":
                return None, "unsupported_xls"
            logger.info(f"[decrypt_try] {p} with password")
            with open(p, "rb") as f:
                office = msoffcrypto.OfficeFile(f)
                office.load_key(password=PASSWORD)
                buf = io.BytesIO()
                office.decrypt(buf)
                buf.seek(0)
                wb = load_workbook(filename=buf, data_only=True, read_only=False, keep_links=False)
                logger.info(f"[decrypt_ok] {p}")
                return wb, "decrypted"
        except PermissionError:
            logger.warning(f"[decrypt_locked] Permission denied: {p}")
            return None, "permission_denied"
        except Exception as e_dec:
            logger.warning(f"[decrypt_failed] {p} :: {e_dec}")
            return None, f"open_error:{e_plain}"

def find_story_id_in_cells(wb, p: Path, mode: str):
    try:
        for name in wb.sheetnames:
            try:
                ws = wb[name]
                logger.debug(f"[sheet_read_start] {p} :: {name}")
                grid = sheet_to_2d_resolved(ws)
                for row in grid:
                    for cell in row:
                        m = RE_STORY.search(normalize_spaces(cell))
                        if m:
                            sid = m.group(0).upper()
                            logger.info(f"[story_found_file] {sid} :: {p} :: {name} :: {mode}")
                            return sid
                logger.debug(f"[sheet_read_done] {p} :: {name}")
            except Exception as se:
                logger.warning(f"[sheet_error] {p} :: {name} :: {se}")
        return None
    except Exception as e:
        logger.error(f"[read_error] {p} :: {e}")
        return None
    finally:
        try:
            wb.close()
        except Exception:
            pass

def find_story_id_in_path(p: Path):
    full = str(p.resolve())
    m = RE_STORY.search(full)
    if m:
        sid = m.group(0).upper()
        logger.info(f"[story_found_path] {sid} :: {p}")
        return sid
    return None

def find_story_id_in_file_or_path(p: Path):
    wb_info = open_workbook_any(p)
    if wb_info and wb_info[0] is not None:
        sid = find_story_id_in_cells(wb_info[0], p, wb_info[1])
        if sid:
            return sid, None
    # fallback to path
    sid_path = find_story_id_in_path(p)
    if sid_path:
        return sid_path, None
    # failure
    reason = wb_info[1] if (wb_info and wb_info[0] is None and len(wb_info) > 1) else "not_found"
    return None, reason

def read_roots_from_file(file_path: Path):
    try:
        lines = file_path.read_text(encoding="utf-8-sig").splitlines()
        logger.info(f"[folders_file_opened] {file_path}")
    except PermissionError:
        logger.error(f"[folders_file_permission] {file_path}"); return []
    except FileNotFoundError:
        logger.error(f"[folders_file_missing] {file_path}"); return []
    roots=[]
    for line in lines:
        s=line.strip().strip('"').strip("'")
        if not s or s.startswith("#"): continue
        p=Path(s)
        if p.is_dir():
            roots.append(p)
            logger.info(f"[root_loaded] {p}")
        else:
            logger.warning(f"[skip_invalid_folder] {s}")
    logger.info(f"[folders_loaded] count={len(roots)}")
    return roots

def iter_excel_files_recursive(root: Path):
    exts = (".xlsx",".xlsm",".xls")
    for base, _, files in os.walk(root):
        for name in files:
            if name.lower().endswith(exts):
                fp = Path(base) / name
                if file_name_matches_final_results(fp):
                    yield fp
                else:
                    logger.debug(f"[file_skipped_by_name] {fp}")

# ==== MAIN ====
def main():
    for f in (OUT_JSONL, FAILED_LIST):
        try:
            if f.exists():
                f.unlink()
                logger.debug(f"[outfile_deleted] {f}")
        except Exception as e:
            logger.warning(f"[outfile_delete_error] {f} :: {e}")

    if not FOLDERS_FILE.exists():
        logger.error(f"[folders_txt_not_found] {FOLDERS_FILE}")
        print("folders.txt not found. Check FOLDERS_FILE in the script."); return

    roots = read_roots_from_file(FOLDERS_FILE)
    if not roots:
        print("No valid parent folders found."); return

    with OUT_JSONL.open("w", encoding="utf-8") as okf, FAILED_LIST.open("w", encoding="utf-8") as ff:
        logger.info(f"[write_open] {OUT_JSONL}")
        logger.info(f"[write_open] {FAILED_LIST}")

        for root in roots:
            root_abs = str(root.resolve())
            logger.info(f"[scan_root] {root_abs}")
            for fp in iter_excel_files_recursive(root):
                sid, err = find_story_id_in_file_or_path(fp)
                if sid:
                    rec = {"root_folder": root_abs, "file_path": str(fp.resolve()), "story_id": sid}
                    try:
                        okf.write(json.dumps(rec, ensure_ascii=False) + "\n")
                        okf.flush()
                        logger.info(f"[write_jsonl] ok :: {fp}")
                    except Exception as we:
                        logger.error(f"[write_error] {OUT_JSONL} :: {we}")
                else:
                    try:
                        ff.write(str(fp.resolve()) + (f"  # {err}" if err else "") + "\n")
                        ff.flush()
                        logger.info(f"[write_failed] {fp} :: {err}")
                    except Exception as fe:
                        logger.error(f"[failed_list_write_error] {FAILED_LIST} :: {fe}")

    logger.info(f"[done] results={OUT_JSONL} failed={FAILED_LIST}")
    print(f"\nDone.\n  ✓ Results: {OUT_JSONL}\n  ✗ Failed:  {FAILED_LIST}\n  🗒️ Log:    {LOG_FILE}")

if __name__ == "__main__":
    main()